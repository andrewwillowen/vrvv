import csv
from dataclasses import replace
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from vrvv.core.export import export_standard_data, export_standard_data_excel
from vrvv.core.units import WAVENUMBER_TO_HZ
from vrvv.ingest.cfour.normalize import normalize_cfour_data
from vrvv.ingest.cfour.parser import CFOURParser

FIXTURE_DIR = Path(__file__).resolve().parents[1] / "fixtures" / "cfour"


def test_export_standard_data_writes_component_csv_files(tmp_path) -> None:
    data = normalize_cfour_data(CFOURParser().parse_raw(FIXTURE_DIR))

    paths = export_standard_data(data, tmp_path / "csv")

    expected_shapes = {
        "equilibrium_rotational_constants.csv": (
            ["rotational_axis", "value_hz"],
            3,
        ),
        "harmonic_frequencies.csv": (
            ["mode_index_zero_based", "frequency_hz"],
            data.n_modes,
        ),
        "cubic_force_constants.csv": (
            [
                "mode_i_zero_based",
                "mode_j_zero_based",
                "mode_k_zero_based",
                "value_hz",
            ],
            data.n_modes**3,
        ),
        "inertial_derivatives.csv": (
            ["mode_index_zero_based", "axis_alpha", "axis_beta", "value_kg^0.5_m"],
            data.n_modes * 3**2,
        ),
        "rotational_derivatives.csv": (
            ["mode_index_zero_based", "axis_alpha", "axis_beta", "value_hz"],
            data.n_modes * 3**2,
        ),
        "coriolis_zetas.csv": (
            [
                "mode_i_zero_based",
                "mode_j_zero_based",
                "rotational_axis",
                "value_dimensionless",
            ],
            data.n_modes**2 * 3,
        ),
        "metadata.csv": (["key", "value"], 1 + len(data.metadata)),
    }

    assert {path.name for path in paths} == set(expected_shapes)
    for name, (headers, row_count) in expected_shapes.items():
        with (tmp_path / "csv" / name).open(newline="", encoding="utf-8") as stream:
            rows = list(csv.reader(stream))
        assert rows[0] == headers
        assert len(rows) == row_count + 1
        assert all(row for row in rows[1:])


def test_export_standard_data_excel_writes_component_worksheets(tmp_path) -> None:
    data = normalize_cfour_data(CFOURParser().parse_raw(FIXTURE_DIR))
    output_path = tmp_path / "standard_data.xlsx"

    path = export_standard_data_excel(data, output_path)

    assert path == output_path
    workbook = load_workbook(output_path, data_only=True)
    expected_headers = {
        "equilibrium_rotational_consts": ["rotational_axis", "value_hz"],
        "harmonic_frequencies": ["mode_index_zero_based", "frequency_hz"],
        "cubic_force_constants": [
            "mode_i_zero_based",
            "mode_j_zero_based",
            "mode_k_zero_based",
            "value_hz",
        ],
        "inertial_derivatives": [
            "mode_index_zero_based",
            "axis_alpha",
            "axis_beta",
            "value_kg^0.5_m",
        ],
        "rotational_derivatives": [
            "mode_index_zero_based",
            "axis_alpha",
            "axis_beta",
            "value_hz",
        ],
        "coriolis_zetas": [
            "mode_i_zero_based",
            "mode_j_zero_based",
            "rotational_axis",
            "value_dimensionless",
        ],
        "metadata": ["key", "value"],
    }
    assert workbook.sheetnames == list(expected_headers)
    for worksheet_name, headers in expected_headers.items():
        worksheet = workbook[worksheet_name]
        assert list(next(worksheet.values)) == headers
    assert workbook["harmonic_frequencies"].max_row == data.n_modes + 1
    assert workbook["cubic_force_constants"].max_row == data.n_modes**3 + 1
    assert workbook["metadata"]["B3"].value == str(FIXTURE_DIR)


def test_export_standard_data_dat_writes_legacy_fortran_records(tmp_path) -> None:
    data = normalize_cfour_data(CFOURParser().parse_raw(FIXTURE_DIR))
    output_path = tmp_path / "standard_data.dat"

    path = data.to_dat(output_path)

    lines = path.read_text(encoding="ascii").splitlines()
    assert path == output_path
    assert len(lines) == 69
    assert lines[0] == "   4   6   1   2"
    assert lines[1].rstrip() == "vrvv StandardData export"
    assert lines[2].rstrip() == "CFOUR"
    assert all(len(line) == 76 for line in lines[1:7])
    np.testing.assert_allclose(
        [float(lines[7][index : index + 12]) for index in range(0, 72, 12)],
        data.harmonic_frequencies.values / WAVENUMBER_TO_HZ,
        atol=5e-7,
    )
    np.testing.assert_allclose(
        [float(lines[8][index : index + 12]) for index in range(0, 36, 12)],
        data.equilibrium_rotational_constants.values / WAVENUMBER_TO_HZ,
        atol=5e-7,
    )
    first_cubic = float(lines[9][:12])
    assert first_cubic == pytest.approx(
        data.cubic_force_constants.values[0, 0, 0] / WAVENUMBER_TO_HZ,
        abs=5e-7,
    )
    c_aa_start = 63
    np.testing.assert_allclose(
        [float(lines[c_aa_start][index : index + 12]) for index in range(0, 72, 12)],
        -data.rotational_derivatives.values[:, 0, 0] / data.harmonic_frequencies.values,
        atol=5e-7,
    )


def test_export_standard_data_dat_honors_metadata_and_rejects_invalid_modes(
    tmp_path,
) -> None:
    data = normalize_cfour_data(CFOURParser().parse_raw(FIXTURE_DIR))
    data = replace(
        data,
        metadata={
            **data.metadata,
            "dat_title": "Custom title",
            "dat_resonance_a_mode": 2,
            "dat_resonance_b_mode": 3,
        },
    )

    path = data.to_dat(tmp_path / "custom.dat")

    assert path.read_text(encoding="ascii").splitlines()[0] == "   4   6   2   3"
    assert path.read_text(encoding="ascii").splitlines()[1].rstrip() == "Custom title"
    with pytest.raises(ValueError, match="must be between"):
        replace(
            data,
            metadata={**data.metadata, "dat_resonance_a_mode": 7},
        ).to_dat(tmp_path / "invalid.dat")
